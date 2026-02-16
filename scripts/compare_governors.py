#!/usr/bin/env python3
"""
Compare governor party data between the Trifectas spreadsheet and the
Supabase elections database (seat_terms table).

Read-only — does not modify any data.
"""

import json
import subprocess
import sys
import time

import openpyxl

# ── Configuration ──────────────────────────────────────────────────────────
XLSX_PATH = "/home/billkramer/Downloads/Trifectas - Political Breakdown Timeline.xlsx"
SUPABASE_REF = "pikcvwulzfxgwfcfssxc"
SUPABASE_TOKEN = "sbp_134edd259126b21a7fc11c7a13c0c8c6834d7fa7"
API_URL = f"https://api.supabase.com/v1/projects/{SUPABASE_REF}/database/query"

PARTY_MAP = {"D": "Democrat", "R": "Republican", "I": "Independent",
             "NP": "Independent", "L": "Independent"}


# ── 1. Read the spreadsheet ───────────────────────────────────────────────
def read_spreadsheet():
    """Return dict: {(state_name, year): party}"""
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb["Govs"]

    # Row 1 has year headers (float) starting in column B
    col_to_year = {}
    for col in range(2, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v is not None:
            col_to_year[col] = int(float(v))

    data = {}
    for row in range(5, 55):  # rows 5-54 = 50 states
        state = ws.cell(row=row, column=1).value
        if not state:
            continue
        state = state.strip()
        for col, year in col_to_year.items():
            party = ws.cell(row=row, column=col).value
            if party and str(party).strip():
                data[(state, year)] = str(party).strip()
    return data


# ── 2. Query the database ────────────────────────────────────────────────
def query_supabase(sql, max_retries=5):
    """Execute SQL via Supabase Management API (curl) with retries."""
    for attempt in range(1, max_retries + 1):
        result = subprocess.run(
            [
                "curl", "-s", "-w", "\n%{http_code}", "-X", "POST",
                API_URL,
                "-H", f"Authorization: Bearer {SUPABASE_TOKEN}",
                "-H", "Content-Type: application/json",
                "-d", json.dumps({"query": sql}),
            ],
            capture_output=True, text=True, timeout=120,
        )
        # Last line of output is the HTTP status code
        lines = result.stdout.rsplit("\n", 1)
        body = lines[0] if len(lines) > 1 else result.stdout
        status = int(lines[1]) if len(lines) > 1 else 0

        if status in (200, 201):
            data = json.loads(body)
            if isinstance(data, dict) and "message" in data:
                print(f"SQL error: {data['message']}", file=sys.stderr)
                sys.exit(1)
            return data
        elif status == 429 and attempt < max_retries:
            wait = 5 * attempt
            print(f"  Rate limited (429). Waiting {wait}s... (attempt {attempt}/{max_retries})")
            time.sleep(wait)
        else:
            print(f"HTTP {status}: {body[:500]}", file=sys.stderr)
            sys.exit(1)


def read_database():
    """Return dict: {(state_name, year): party}"""
    sql = """
    SELECT s.state_name, s.abbreviation,
           st.start_date, st.end_date,
           c.full_name, st.party, st.caucus
    FROM seat_terms st
    JOIN seats se ON st.seat_id = se.id
    JOIN districts d ON se.district_id = d.id
    JOIN states s ON d.state_id = s.id
    JOIN candidates c ON st.candidate_id = c.id
    WHERE se.office_type = 'Governor'
    ORDER BY s.abbreviation, st.start_date
    """
    print("Querying Supabase for governor seat_terms…")
    rows = query_supabase(sql)
    print(f"  Got {len(rows)} seat_term records.")

    data = {}
    for r in rows:
        state = r["state_name"]
        start_year = int(r["start_date"][:4])
        # end_date may be null for current office-holders
        if r["end_date"]:
            end_year = int(r["end_date"][:4])
        else:
            end_year = 2025  # assume current through 2025

        # Use caucus if available, else party
        raw = r["caucus"] if r["caucus"] else r["party"]
        party = PARTY_MAP.get(raw, raw)

        # Assign to every year the term overlaps
        # A term from e.g. 2019-01-14 to 2023-01-09 covers years 2019-2022
        # But inauguration years count: if start is Jan 2019, that year counts.
        # If end is Jan 2023, the governor was in office briefly in 2023 too,
        # but the *new* governor takes over that year, so we include end_year
        # only if the end_date is after Jan 15 (rough midpoint of inauguration).
        # Simpler approach: include start_year through end_year-1, plus end_year
        # if end_date is after June 30 (i.e. they served most of the year).
        # Actually, the spreadsheet records who was governor for the *majority*
        # of the year. Inauguration is typically early January, so:
        # - start_year: always include (they started that year)
        # - end_year: include only if end_date >= July 1 of that year
        #   (meaning they served at least half the year)
        # For null end_date (current), include through 2025.

        if r["end_date"]:
            end_month = int(r["end_date"][5:7])
            end_day = int(r["end_date"][8:10])
            # If term ends after June 30, include that year
            if end_month > 6 or (end_month == 6 and end_day >= 30):
                last_year = end_year
            else:
                last_year = end_year - 1
        else:
            last_year = 2025

        for y in range(start_year, last_year + 1):
            # If there's already an entry for this state-year from a later
            # term, the later term wins (they iterate in start_date order)
            data[(state, y)] = party

    return data, rows


# ── 3. Compare ────────────────────────────────────────────────────────────
def compare(sheet_data, db_data):
    all_keys = set(sheet_data.keys()) | set(db_data.keys())
    # Only compare for years 1970-2025 (spreadsheet range)
    all_keys = {(s, y) for s, y in all_keys if 1970 <= y <= 2025}

    disagreements = []
    sheet_only = []
    db_only = []

    for key in sorted(all_keys):
        state, year = key
        in_sheet = key in sheet_data
        in_db = key in db_data

        if in_sheet and in_db:
            sp = sheet_data[key]
            dp = db_data[key]
            if sp != dp:
                disagreements.append((state, year, sp, dp))
        elif in_sheet and not in_db:
            sheet_only.append((state, year, sheet_data[key]))
        elif in_db and not in_sheet:
            db_only.append((state, year, db_data[key]))

    return disagreements, sheet_only, db_only


# ── 4. Report ─────────────────────────────────────────────────────────────
def report(disagreements, sheet_only, db_only, db_rows):
    print("\n" + "=" * 80)
    print("GOVERNOR PARTY COMPARISON: Spreadsheet vs. Database")
    print("=" * 80)

    # Build a lookup of DB governor names by state+year for context
    db_names = {}
    for r in db_rows:
        state = r["state_name"]
        start_year = int(r["start_date"][:4])
        end_year = int(r["end_date"][:4]) if r["end_date"] else 2025
        for y in range(start_year, end_year + 1):
            db_names[(state, y)] = r["full_name"]

    # --- Disagreements ---
    print(f"\n{'─' * 80}")
    print(f"1. DISAGREEMENTS (both sources have data but differ): {len(disagreements)}")
    print(f"{'─' * 80}")
    if disagreements:
        # Group by state
        from collections import defaultdict
        by_state = defaultdict(list)
        for state, year, sp, dp in disagreements:
            gov_name = db_names.get((state, year), "?")
            by_state[state].append((year, sp, dp, gov_name))

        for state in sorted(by_state):
            items = by_state[state]
            print(f"\n  {state}:")
            for year, sp, dp, gov in items:
                print(f"    {year}: Sheet={sp:12s}  DB={dp:12s}  (DB gov: {gov})")
    else:
        print("  None — perfect agreement!")

    # --- Sheet only (gaps in DB) ---
    print(f"\n{'─' * 80}")
    print(f"2. SPREADSHEET ONLY (gaps in DB): {len(sheet_only)} state-year entries")
    print(f"{'─' * 80}")
    if sheet_only:
        from collections import defaultdict
        by_state = defaultdict(list)
        for state, year, party in sheet_only:
            by_state[state].append((year, party))

        for state in sorted(by_state):
            items = sorted(by_state[state])
            years = [i[0] for i in items]
            # Compress into ranges
            ranges = compress_years(years)
            parties = set(i[1] for i in items)
            print(f"  {state}: {ranges} ({', '.join(sorted(parties))})")
    else:
        print("  None — DB covers all spreadsheet entries!")

    # --- DB only ---
    print(f"\n{'─' * 80}")
    print(f"3. DB ONLY (not in spreadsheet, years outside 1970-2025): {len(db_only)} entries")
    print(f"{'─' * 80}")
    if db_only:
        from collections import defaultdict
        by_state = defaultdict(list)
        for state, year, party in db_only:
            by_state[state].append((year, party))
        for state in sorted(by_state):
            items = sorted(by_state[state])
            years = [i[0] for i in items]
            ranges = compress_years(years)
            print(f"  {state}: {ranges}")
    else:
        print("  None")

    # --- Summary stats ---
    all_sheet_states = sorted(set(s for s, y in sheet_data))
    all_db_states = sorted(set(s for s, y in db_data))
    sheet_years = sorted(set(y for s, y in sheet_data))
    db_years = sorted(set(y for s, y in db_data))

    print(f"\n{'─' * 80}")
    print("SUMMARY")
    print(f"{'─' * 80}")
    print(f"  Spreadsheet: {len(all_sheet_states)} states, years {min(sheet_years)}-{max(sheet_years)}, {len(sheet_data)} total entries")
    print(f"  Database:    {len(all_db_states)} states, years {min(db_years)}-{max(db_years)}, {len(db_data)} total entries")
    overlap = set(sheet_data.keys()) & set(db_data.keys())
    agree = len(overlap) - len(disagreements)
    print(f"  Overlap:     {len(overlap)} state-year pairs compared")
    print(f"  Agreement:   {agree} ({100*agree/len(overlap):.1f}%)")
    print(f"  Disagreements: {len(disagreements)}")
    print(f"  Sheet only:  {len(sheet_only)} (DB gaps)")
    print(f"  DB only:     {len(db_only)} (outside sheet range)")


def compress_years(years):
    """Turn [1970,1971,1972,1975,1976] into '1970-1972, 1975-1976'."""
    if not years:
        return ""
    years = sorted(years)
    ranges = []
    start = years[0]
    end = years[0]
    for y in years[1:]:
        if y == end + 1:
            end = y
        else:
            ranges.append(f"{start}-{end}" if start != end else str(start))
            start = end = y
    ranges.append(f"{start}-{end}" if start != end else str(start))
    return ", ".join(ranges)


# ── Main ──────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("Reading spreadsheet…")
    sheet_data = read_spreadsheet()
    print(f"  Got {len(sheet_data)} state-year entries from spreadsheet.")

    db_data, db_rows = read_database()
    print(f"  Got {len(db_data)} state-year entries from database.")

    disagreements, sheet_only, db_only = compare(sheet_data, db_data)
    report(disagreements, sheet_only, db_only, db_rows)
