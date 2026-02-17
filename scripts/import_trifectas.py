#!/usr/bin/env python3
"""
Import trifecta, legislative control, and governor party data from
'Trifectas - Political Breakdown Timeline.xlsx' into the trifectas table.

Reads all three sheets:
- Trifectas: trifecta_status (Republican/Democrat/Split) by state/year (1992-2026)
- Leg: legislature_status (Republican/Democrat/Split) by state/year (2000-2025)
- Govs: governor_party (Republican/Democrat/Independent) by state/year (1970-2025)

Usage:
    python3 scripts/import_trifectas.py [--dry-run]
"""

import sys
import json
import os
import time
import requests
import openpyxl
import sys as _sys, os as _os
_sys.path.insert(0, _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), '..'))
from db_config import TOKEN, PROJECT_REF, API_URL

XLSX_PATH = os.path.expanduser("~/Downloads/Trifectas - Political Breakdown Timeline.xlsx")


def run_sql(query, attempt=1, max_attempts=5):
    """Execute SQL via Management API with retry logic."""
    headers = {
        'Authorization': f'Bearer {TOKEN}',
        'Content-Type': 'application/json'
    }
    resp = requests.post(API_URL, headers=headers, json={'query': query})
    if resp.status_code == 429 and attempt < max_attempts:
        wait = 5 * attempt
        print(f"  Rate limited, waiting {wait}s (attempt {attempt}/{max_attempts})...")
        time.sleep(wait)
        return run_sql(query, attempt + 1, max_attempts)
    if resp.status_code not in (200, 201):
        print(f"  SQL error ({resp.status_code}): {resp.text[:200]}")
        return None
    return resp.json()


def read_sheet_data(wb, sheet_name):
    """Read a sheet into a dict of {state_name: {year: value}}."""
    ws = wb[sheet_name]
    data = {}

    # Row 1 is headers: col A = "State", col B onward = years
    headers = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        headers.append(val)

    # Parse year columns (skip col A which is state name)
    year_cols = {}  # {col_index: year}
    for col_idx, header in enumerate(headers):
        if header is None:
            continue
        try:
            year = int(header)
            if 1960 <= year <= 2030:
                year_cols[col_idx] = year
        except (ValueError, TypeError):
            continue

    # Rows 2-4 are summary rows (R Totals, D Totals, S/I Totals)
    # Rows 5-54 are state data (50 states)
    for row in range(5, 55):
        state_name = ws.cell(row=row, column=1).value
        if not state_name or not isinstance(state_name, str):
            continue
        state_name = state_name.strip()
        data[state_name] = {}
        for col_idx, year in year_cols.items():
            val = ws.cell(row=row, column=col_idx + 1).value  # openpyxl is 1-indexed
            if val and isinstance(val, str):
                val = val.strip()
                if val in ('Republican', 'Democrat', 'Independent', 'Split'):
                    data[state_name][year] = val

    return data


def main():
    dry_run = '--dry-run' in sys.argv

    print(f"Reading {XLSX_PATH}...")
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    print("\nParsing sheets...")
    trifecta_data = read_sheet_data(wb, 'Trifectas')
    leg_data = read_sheet_data(wb, 'Leg')
    gov_data = read_sheet_data(wb, 'Govs')
    wb.close()

    # Get state name -> id mapping from DB
    print("\nFetching state IDs...")
    result = run_sql("SELECT id, state_name FROM states ORDER BY id")
    if not result:
        print("ERROR: Could not fetch states")
        sys.exit(1)
    state_map = {row['state_name']: row['id'] for row in result}

    # Build combined records
    # For each state/year, combine data from all three sheets
    records = {}  # {(state_name, year): {governor_party, legislature_status, trifecta_status}}

    # First pass: trifecta data (1992-2026)
    for state, years in trifecta_data.items():
        for year, status in years.items():
            records[(state, year)] = {'trifecta_status': status}

    # Add legislature_status from Leg sheet
    for state, years in leg_data.items():
        for year, status in years.items():
            key = (state, year)
            if key not in records:
                records[key] = {}
            records[key]['legislature_status'] = status

    # Add governor_party from Govs sheet
    for state, years in gov_data.items():
        for year, status in years.items():
            key = (state, year)
            if key not in records:
                records[key] = {}
            records[key]['governor_party'] = status

    # For years that have gov + leg but no trifecta, compute it
    for key, rec in records.items():
        if 'trifecta_status' not in rec:
            gov = rec.get('governor_party')
            leg = rec.get('legislature_status')
            if gov and leg:
                if gov == leg:
                    rec['trifecta_status'] = gov
                else:
                    rec['trifecta_status'] = 'Split'

    # Report Split legislatures for investigation
    split_legs = [(state, year) for (state, year), rec in records.items()
                  if rec.get('legislature_status') == 'Split']
    split_legs.sort(key=lambda x: (x[1], x[0]))

    print(f"\nFound {len(split_legs)} Split legislature instances:")
    for state, year in split_legs:
        rec = records[(state, year)]
        gov = rec.get('governor_party', '?')
        tri = rec.get('trifecta_status', '?')
        print(f"  {year} {state}: Leg=Split, Gov={gov}, Trifecta={tri}")

    # Filter to records that have trifecta_status
    valid_records = {k: v for k, v in records.items() if 'trifecta_status' in v}
    print(f"\nTotal records to import: {len(valid_records)}")

    # Summary stats
    years_covered = sorted(set(y for _, y in valid_records.keys()))
    print(f"Years covered: {min(years_covered)}-{max(years_covered)}")

    # Count by trifecta status for 2025
    for check_year in [2025, 2026]:
        year_recs = {k: v for k, v in valid_records.items() if k[1] == check_year}
        if year_recs:
            r = sum(1 for v in year_recs.values() if v['trifecta_status'] == 'Republican')
            d = sum(1 for v in year_recs.values() if v['trifecta_status'] == 'Democrat')
            s = sum(1 for v in year_recs.values() if v['trifecta_status'] == 'Split')
            print(f"  {check_year}: {r}R, {d}D, {s}S (total {r+d+s})")

    if dry_run:
        print("\n[DRY RUN] Would insert records. Exiting.")
        return

    # Batch insert
    print("\nInserting into trifectas table...")
    batch_size = 50
    sorted_records = sorted(valid_records.items(), key=lambda x: (x[0][1], x[0][0]))

    inserted = 0
    for i in range(0, len(sorted_records), batch_size):
        batch = sorted_records[i:i + batch_size]
        values = []
        for (state, year), rec in batch:
            state_id = state_map.get(state)
            if not state_id:
                print(f"  WARNING: No state_id for '{state}', skipping")
                continue
            gov = rec.get('governor_party')
            leg = rec.get('legislature_status')
            tri = rec['trifecta_status']
            gov_sql = f"'{gov}'" if gov else 'NULL'
            leg_sql = f"'{leg}'" if leg else 'NULL'
            values.append(f"({state_id}, {year}, {gov_sql}, {leg_sql}, '{tri}')")

        if not values:
            continue

        sql = f"""INSERT INTO trifectas (state_id, year, governor_party, legislature_status, trifecta_status)
VALUES {', '.join(values)}
ON CONFLICT (state_id, year) DO UPDATE SET
    governor_party = EXCLUDED.governor_party,
    legislature_status = EXCLUDED.legislature_status,
    trifecta_status = EXCLUDED.trifecta_status"""

        result = run_sql(sql)
        if result is not None:
            inserted += len(values)
            print(f"  Inserted batch {i // batch_size + 1} ({len(values)} records)")
        else:
            print(f"  FAILED batch {i // batch_size + 1}")

    print(f"\nDone! Inserted {inserted} trifecta records.")


if __name__ == '__main__':
    main()
