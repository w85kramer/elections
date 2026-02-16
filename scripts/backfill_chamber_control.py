#!/usr/bin/env python3
"""
Backfill historical chamber_control data from spreadsheets.

Sources:
- 'Comparison 2022 vs 2020 Elections Data.xlsx' Legislatures sheet:
  - "2021" columns = post-2020 election composition → effective_date 2021-01-01
  - "2023" columns = post-2022 election composition → effective_date 2023-01-01

- 'Comparison 2024 vs 2022 Elections Data.xlsx' Legislatures sheet:
  - "2024" columns = pre-2024 election (same as post-2022, skip as duplicate of 2023)
  - "2025" columns = post-2024 election (we already have this, skip)

Usage:
    python3 scripts/backfill_chamber_control.py [--dry-run]
"""

import sys
import os
import time
import requests
import openpyxl

DRY_RUN = '--dry-run' in sys.argv

FILE_2022 = os.path.expanduser("~/Downloads/Comparison 2022 vs 2020 Elections Data.xlsx")

ENV_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), '.env')
def load_env():
    env = {}
    with open(ENV_PATH) as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith('#') and '=' in line:
                k, v = line.split('=', 1)
                env[k.strip()] = v.strip()
    return env

env = load_env()
TOKEN = env['SUPABASE_MANAGEMENT_TOKEN']
API_URL = f'https://api.supabase.com/v1/projects/pikcvwulzfxgwfcfssxc/database/query'


def run_sql(query, attempt=1):
    headers = {'Authorization': f'Bearer {TOKEN}', 'Content-Type': 'application/json'}
    resp = requests.post(API_URL, headers=headers, json={'query': query})
    if resp.status_code == 429 and attempt < 5:
        time.sleep(5 * attempt)
        return run_sql(query, attempt + 1)
    if resp.status_code not in (200, 201):
        print(f"  SQL error ({resp.status_code}): {resp.text[:200]}")
        return None
    return resp.json()


def determine_control(d, r, other):
    """Determine control_status from seat counts."""
    total = d + r + other
    if total == 0:
        return 'Nonpartisan'
    majority = total // 2 + 1
    if d >= majority:
        return 'D'
    elif r >= majority:
        return 'R'
    else:
        return 'Coalition'  # No party has majority


def read_legislatures_sheet(filepath):
    """Read the Legislatures sheet and extract seat count snapshots."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['Legislatures']

    # Find column indices from header row
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val:
            val = str(val).strip()
            # Track all headers with their column index
            if val not in headers:
                headers[val] = col
            else:
                # Duplicate header - this is the "post" set
                headers[val + '_post'] = col

    # The layout is:
    # A=State, B=Chamber, C=Dist.Type, D=Total Seats,
    # E=pre(D), F=pre(R), G=pre(I), H=pre(V), I=PreMaj%,
    # J=Total Seats, K=post(D), L=post(R), M=post(I), N=post(V), O=PostMaj%
    # For the 2022 file: pre=2021, post=2023

    snapshots = {'pre': [], 'post': []}

    # Row 2 is totals, rows 3+ are state data
    for row in range(3, ws.max_row + 1):
        state = ws.cell(row=row, column=1).value
        chamber = ws.cell(row=row, column=2).value
        if not state or not chamber:
            continue
        state = str(state).strip()
        chamber = str(chamber).strip()
        if len(state) > 2:  # Skip non-state rows
            continue

        total_pre = ws.cell(row=row, column=4).value or 0
        pre_d = ws.cell(row=row, column=5).value or 0
        pre_r = ws.cell(row=row, column=6).value or 0
        pre_i = ws.cell(row=row, column=7).value or 0
        pre_v = ws.cell(row=row, column=8).value or 0

        total_post = ws.cell(row=row, column=10).value or 0
        post_d = ws.cell(row=row, column=11).value or 0
        post_r = ws.cell(row=row, column=12).value or 0
        post_i = ws.cell(row=row, column=13).value or 0
        post_v = ws.cell(row=row, column=14).value or 0

        # Convert to int safely
        def to_int(v):
            try:
                return int(v)
            except (TypeError, ValueError):
                return 0

        snapshots['pre'].append({
            'state': state, 'chamber': chamber,
            'total': to_int(total_pre),
            'd': to_int(pre_d), 'r': to_int(pre_r),
            'other': to_int(pre_i), 'vacant': to_int(pre_v)
        })
        snapshots['post'].append({
            'state': state, 'chamber': chamber,
            'total': to_int(total_post),
            'd': to_int(post_d), 'r': to_int(post_r),
            'other': to_int(post_i), 'vacant': to_int(post_v)
        })

    wb.close()
    return snapshots


def main():
    # Get state abbrev -> id mapping
    print("Fetching state IDs...")
    result = run_sql("SELECT id, abbreviation FROM states")
    state_map = {r['abbreviation']: r['id'] for r in result}

    # Map chamber names
    def map_chamber(chamber):
        mapping = {
            'House': 'House',
            'Senate': 'Senate',
            'Unicameral': 'Legislature',
            'Assembly': 'House',
            'House of Delegates': 'House',
        }
        return mapping.get(chamber, chamber)

    # Read 2022 file (has 2021 pre and 2023 post snapshots)
    print(f"\nReading {FILE_2022}...")
    snapshots = read_legislatures_sheet(FILE_2022)

    print(f"  Pre (2021): {len(snapshots['pre'])} chambers")
    print(f"  Post (2023): {len(snapshots['post'])} chambers")

    # Build insert statements for 2021-01-01 and 2023-01-01
    all_values = []

    for effective_date, data in [('2021-01-01', snapshots['pre']), ('2023-01-01', snapshots['post'])]:
        for rec in data:
            state_id = state_map.get(rec['state'])
            if not state_id:
                print(f"  WARNING: no state_id for '{rec['state']}'")
                continue
            chamber = map_chamber(rec['chamber'])
            control = determine_control(rec['d'], rec['r'], rec['other'])
            total = rec['total']
            majority_threshold = total // 2 + 1

            all_values.append(
                f"({state_id}, '{chamber}', '{effective_date}', '{control}', "
                f"{rec['d']}, {rec['r']}, {rec['other']}, {rec['vacant']}, "
                f"{total}, {majority_threshold})"
            )

    print(f"\nTotal records to insert: {len(all_values)}")

    # Count by date
    count_2021 = sum(1 for v in all_values if "'2021-01-01'" in v)
    count_2023 = sum(1 for v in all_values if "'2023-01-01'" in v)
    print(f"  2021-01-01: {count_2021} chambers")
    print(f"  2023-01-01: {count_2023} chambers")

    if DRY_RUN:
        print("\n[DRY RUN] Would insert records. Exiting.")
        return

    # Insert in batches
    print("\nInserting...")
    inserted = 0
    batch_size = 50
    for i in range(0, len(all_values), batch_size):
        batch = all_values[i:i + batch_size]
        sql = f"""INSERT INTO chamber_control
            (state_id, chamber, effective_date, control_status,
             d_seats, r_seats, other_seats, vacant_seats, total_seats, majority_threshold)
        VALUES {', '.join(batch)}
        ON CONFLICT (state_id, chamber, effective_date) DO UPDATE SET
            control_status = EXCLUDED.control_status,
            d_seats = EXCLUDED.d_seats,
            r_seats = EXCLUDED.r_seats,
            other_seats = EXCLUDED.other_seats,
            vacant_seats = EXCLUDED.vacant_seats,
            total_seats = EXCLUDED.total_seats,
            majority_threshold = EXCLUDED.majority_threshold"""

        result = run_sql(sql)
        if result is not None:
            inserted += len(batch)
            print(f"  Batch {i // batch_size + 1}: {len(batch)} records")
        else:
            print(f"  FAILED batch {i // batch_size + 1}")

    print(f"\nDone! Inserted {inserted} chamber_control records.")

    # Summary
    result = run_sql("""
        SELECT effective_date, count(*) as chambers,
            sum(d_seats) as total_d, sum(r_seats) as total_r
        FROM chamber_control
        GROUP BY effective_date
        ORDER BY effective_date
    """)
    if result:
        print("\nChamber control snapshots in DB:")
        for r in result:
            print(f"  {r['effective_date']}: {r['chambers']} chambers, {r['total_d']}D / {r['total_r']}R")


if __name__ == '__main__':
    main()
